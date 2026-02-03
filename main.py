import base64
import csv
import os
import uuid
from pathlib import Path
from typing import Any

import pyodbc
from dotenv import load_dotenv
from openpyxl import load_workbook


# ================= OpenEdge Connection =================

class OpenEdgeConnection:
    def __init__(self) -> None:
        self._load_env()
        self.conn_str = (
            f"DSN={os.getenv('OE_DSN')};"
            f"UID={os.getenv('OE_USER')};"
            f"PWD={os.getenv('OE_PASSWORD')};"
        )

    def _load_env(self) -> None:
        base_dir = Path(__file__).resolve().parent
        load_dotenv(base_dir / "config.env")

    def connect(self):
        return pyodbc.connect(self.conn_str)


# ================= Data Processor =================

class DataProcessor:
    def __init__(self, connection: OpenEdgeConnection) -> None:
        self.connection = connection
        os.makedirs("images", exist_ok=True)
        os.makedirs("output", exist_ok=True)

    # -------- STREAMING CSV EXPORT (CORE FIX) --------

    def export_query_to_csv(self, query: str, csv_name: str) -> None:
        csv_path = Path("output") / f"{csv_name}.csv"

        with self.connection.connect() as conn:
            cur = conn.cursor()
            cur.execute(query)

            columns = [col[0] for col in cur.description]

            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(columns)

                while True:
                    rows = cur.fetchmany(5000)  # batch size
                    if not rows:
                        break

                    for row in rows:
                        writer.writerow([
                            self.process_cell(value, columns[idx])
                            for idx, value in enumerate(row)
                        ])

        print(f"✅ CSV created (streamed): {csv_path}")

    # -------- CELL PROCESSING --------

    def process_cell(self, value: Any, column_name: str) -> str:
        if value is None:
            return ""

        if isinstance(value, (bytes, bytearray)):
            if self._is_real_image(value):
                return self._save_image(value, column_name)
            return value.hex()

        if isinstance(value, str) and value.startswith("data:image"):
            try:
                img_bytes = base64.b64decode(value.split(",", 1)[1])
                if self._is_real_image(img_bytes):
                    return self._save_image(img_bytes, column_name)
            except Exception:
                pass

        return str(value)

    def _is_real_image(self, img_bytes: bytes) -> bool:
        return (
            img_bytes.startswith(b"\x89PNG") or
            img_bytes.startswith(b"\xff\xd8\xff") or
            img_bytes.startswith(b"RIFF")
        )

    def _get_image_extension(self, img_bytes: bytes) -> str:
        if img_bytes.startswith(b"\x89PNG"):
            return "png"
        if img_bytes.startswith(b"\xff\xd8\xff"):
            return "jpg"
        if img_bytes.startswith(b"RIFF"):
            return "webp"
        return "bin"

    def _save_image(self, img_bytes: bytes, column_name: str) -> str:
        ext = self._get_image_extension(img_bytes)
        filename = f"{column_name}_{uuid.uuid4().hex}.{ext}"
        path = Path("images") / filename

        with open(path, "wb") as f:
            f.write(img_bytes)

        return str(path)


# ================= Excel Helper =================

def get_excel_file() -> Path:
    base_dir = Path(__file__).resolve().parent

    excel_files = [
        f for f in base_dir.glob("*.xlsx")
        if not f.name.startswith("~$")
    ]

    if not excel_files:
        raise FileNotFoundError(
            "❌ No Excel file found.\n"
            "✔ Place Excel next to this script\n"
            "✔ Close Excel before running"
        )

    return excel_files[0]


# ================= MAIN =================

def main():
    db = OpenEdgeConnection()
    processor = DataProcessor(db)

    excel_file = get_excel_file()
    print(f"📄 Using Excel file: {excel_file.name}")

    wb = load_workbook(excel_file)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        query = sheet[f"C{row}"].value
        csv_name_cell = sheet[f"D{row}"]
        status_cell = sheet[f"E{row}"]

        if not query:
            continue

        if not csv_name_cell.value:
            status_cell.value = "ERROR: CSV name missing"
            continue

        if status_cell.value:
            continue  # already processed

        csv_name = csv_name_cell.value.replace(".csv", "").strip()

        print(f"▶ Processing row {row}: {csv_name}")

        try:
            status_cell.value = "Processing..."
            wb.save(excel_file)

            processor.export_query_to_csv(query, csv_name)

            status_cell.value = "Done"

        except Exception as e:
            status_cell.value = f"ERROR: {str(e)}"

        wb.save(excel_file)

    print("🎯 Excel updated successfully")


if __name__ == "__main__":
    main()
